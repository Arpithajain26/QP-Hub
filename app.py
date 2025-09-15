from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory
import os
import json
import openpyxl
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = "supersecretkey"

# ======== PATH CONFIG =========
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "users.xlsx")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
QP_DB = os.path.join(BASE_DIR, "qp_data.json")
ALLOWED_EXTENSIONS = {"pdf", "doc", "docx"}

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# ======== HELPERS =========
def create_excel_if_not_exists():
    """Create users.xlsx with headers if not already present."""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["Full Name", "Email", "Password"])
        wb.save(EXCEL_FILE)

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def load_qp_data():
    if os.path.exists(QP_DB):
        with open(QP_DB, "r") as f:
            return json.load(f)
    return []

def save_qp_data(data):
    with open(QP_DB, "w") as f:
        json.dump(data, f, indent=4)

# ======== ROUTES =========
@app.route("/")
def home():
    return redirect(url_for("signup"))

@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not first_name or not last_name or not email or not password:
            return render_template("signup.html", error="⚠️ All fields are required.")

        if password != confirm_password:
            return render_template("signup.html", error="❌ Passwords do not match.")

        full_name = f"{first_name} {last_name}"

        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == email:
                    return render_template("signup.html", error="⚠️ User already exists. Try logging in!")

            ws.append([full_name, email, password])
            wb.save(EXCEL_FILE)

            return redirect(url_for("success", name=full_name, source="signup"))

        except Exception as e:
            print(f"Error writing to Excel: {e}")
            return render_template("signup.html", error="❌ Could not save user. Please try again later.")

    return render_template("signup.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")

        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == email and row[2] == password:
                    return redirect(url_for("success", name=row[0], source="login"))

            return render_template("login.html", error="❌ Invalid email or password.")

        except Exception as e:
            print(f"Error reading Excel: {e}")
            return render_template("login.html", error="⚠️ Server error. Try again later.")

    return render_template("login.html")

@app.route("/success")
def success():
    name = request.args.get("name", "User")
    source = request.args.get("source", "signup")
    message = "✅ Signup Successful!" if source == "signup" else "✅ Login Successful!"
    return render_template("success.html", name=name, message=message)

@app.route("/dashboard")
def dashboard():
    name = request.args.get("name", "User")
    return render_template("dashboard.html", name=name)

@app.route("/upload", methods=["GET", "POST"])
def upload():
    if request.method == "POST":
        subject = request.form.get("subject")
        subject_code = request.form.get("subject_code")
        branch = request.form.get("branch")
        semester = request.form.get("semester")
        year = request.form.get("year")
        file = request.files.get("file")

        if not (subject and subject_code and branch and semester and year and file):
            flash("⚠️ All fields are required!")
            return redirect(url_for("upload"))

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)

            qp_data = load_qp_data()
            qp_data.append({
                "subject": subject,
                "subject_code": subject_code,
                "branch": branch,
                "semester": semester,
                "year": year,
                "filename": filename
            })
            save_qp_data(qp_data)

            flash("✅ File uploaded successfully!")
            return redirect(url_for("upload"))
        else:
            flash("❌ Invalid file type. Only PDF, DOC, DOCX allowed.")
            return redirect(url_for("upload"))

    return render_template("upload.html")

@app.route("/view_qps")
def view_qps():
    qp_list = load_qp_data()
    return render_template("view_qps.html", qp_list=qp_list)

@app.route("/download/<filename>")
def download_qp(filename):
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)

# ======== RUN APP =========
if __name__ == "__main__":
    create_excel_if_not_exists()
    app.run(debug=True)
